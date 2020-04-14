import os
from openpyxl import load_workbook

class CommitteeForm:
    form = {
            'committee_due_date': {
                'label': 'B7',
                'value': 'C7',
                'default': '',
            },
            'date_routing_initiated': {
                'label': 'B9',
                'value': 'C9',
                'default': '',
            },
            'type_of_exception_request': {
                'label': 'B9',
                'value': 'C9',
                'default': '',
                'options': [
                    'Full embargo - restrict from campus',
                    'Extend 6-month embargo',
                ],
            },
            'current_embargo_expiration': {
                'label': 'B10',
                'value': 'C10',
                'default': '',
            },
            'thesis_author': {
                'label': 'B14',
                'value': 'C14',
                'default': '',
            },
            'thesis_title': {
                'label': 'B15',
                'value': 'C15',
                'default': '',
            },
            'advisor_name': {
                'label': 'B16',
                'value': 'C16',
                'default': '',
            },
            'division': {
                'label': 'B17',
                'value': 'C17',
                'default': '',
            },
            'grad_year': {
                'label': 'B18',
                'value': 'C18',
                'default': '',
            },
            'requestor_name': {
                'label': 'B22',
                'value': 'C22',
                'default': '',
            },
            'embargo_reason': {
                'label': 'B23',
                'value': 'C23',
                'default': '',
            },
            'embargo_abstract': {
                'label': 'B24',
                'value': 'C24',
                'default': 'No',
                'options': [
                    'Yes',
                    'No',
                ],
            },
    }

    def render(self, xl_filename, name, obj):
        dir_path = os.path.dirname(os.path.realpath(__file__))
        xlt_filename = os.path.join(dir_path, 'templates', 'thesis_embargo_committee.xlsx')
        wb = load_workbook(filename = xlt_filename)
        ws = wb['Form']
        form = self.form
        for key in obj:
            if (key in form) and ('value' in form[key]):
                cell = form[key]['value']
                ws[cell] = obj[key]
        wb.save(xl_filename)
        

